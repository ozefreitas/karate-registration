from .models import Bracket, Match, ScoringEntry, ScoringRound
import draw.serializers as serializers
from core.views import MultipleSerializersMixIn
from core.permissions import IsAuthenticatedOrReadOnly, IsTechnicianOrAdmin, IsTechnicianOrAdminforPOST, IsAdminRoleorHigherForGET, IsAdminRoleorHigher
from registration.serializers.base import CompactPersonSerializer, CompactTeamSerializer
from registration.serializers.serializers import ClassificationsSerializer
from registration.utils.utils import next_power_of_2
from registration.models import Person, Classification, Team
from .filters import BracketsFilters, MatchesFilters, ScoringEntriesFilters
from events.models import EventDorsal, DisciplineTeam, DisciplineMember
from django.db.models import Max
from .utils.draw_utils import seed_registrations_by_club

from django.http import HttpResponse
import io
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

from drf_spectacular.utils import extend_schema

from django.utils import timezone

# Create your views here.


def get_dorsal(person, dorsals):
        if not person:
            return None
        dorsal = dorsals.get(person.id)
        return str(dorsal).zfill(3) if dorsal is not None else None

class BracketViewSet(MultipleSerializersMixIn, viewsets.ModelViewSet):
    queryset=Bracket.objects.all()
    serializer_class=serializers.BracketSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filterset_class = BracketsFilters
    pagination_class = None

    serializer_classes = {
        "create": serializers.CreateBracketSerializer,
    }

    @extend_schema(responses=CompactPersonSerializer(many=True))
    @action(detail=True, methods=["get"], url_path="persons")
    def persons(self, request, pk=None):
        bracket = self.get_object()
        
        persons = Person.objects.filter(
            disciplinemember__category=bracket.category,
            disciplinemember__discipline=bracket.discipline,
        )

        serializer = CompactPersonSerializer(persons, many=True)
        return Response(serializer.data)
    
    @extend_schema(responses=CompactTeamSerializer(many=True))
    @action(detail=True, methods=["get"], url_path="teams")
    def teams(self, request, pk=None):
        bracket = self.get_object()
        
        teams = Team.objects.filter(
            category=bracket.category,
            disciplineteam__discipline=bracket.discipline,
        )

        serializer = CompactTeamSerializer(teams, many=True)
        return Response(serializer.data)

    @action(
        detail=True,
        methods=["post"],
        url_path="generate_bracket_draw",
        permission_classes=[IsAdminRoleorHigher]
    )
    def generate_bracket_draw(self, request, pk=None):
        bracket = self.get_object()
        category = bracket.category
        discipline = bracket.discipline
        bracket_draw_type = bracket.draw_type
        event = bracket.event
        # WORKAROUND!! Need to create a modal after ###
        finals_size = 8
        split_clubs = True

        will_go_to_scoring = False
        if bracket_draw_type == "Misto" and finals_size:
            will_go_to_scoring = True

        if discipline.is_team:
            # Retrieves all teams for the current category
            category_registrations = DisciplineTeam.objects.filter(
                team__category=category,
                discipline=discipline,
            ).order_by("id")

        else:
            # Retrieves all registrations for the current category
            category_registrations = DisciplineMember.objects.filter(
                category=category,
                discipline=discipline
            ).order_by("id")
        
        if split_clubs:
            registrations = seed_registrations_by_club(list(category_registrations), "team" if discipline.is_team else "indiv")
        else:
            registrations = list(category_registrations)
        total_players = len(registrations)

        # Categories with less than 2 registration will not take place, so proceed to the next category
        if len(registrations) < 2:
            return Response(
                    {"message": "Este Escalão não tem inscrições suficientes para proceder a um sorteio."},
                    status=status.HTTP_400_BAD_REQUEST
                 )
        
        bracket_name = f'{discipline.name} {category.name} {category.gender}'
        if category.min_weight is not None:
            bracket_name += f' +{category.min_weight}kg'
        elif category.max_weight is not None:
            bracket_name += f' -{category.max_weight}kg'

        # delete bracket to create new one (will cascade to scoring rounds and entries)
        bracket.delete()

        new_bracket = Bracket.objects.create(
                                            name=bracket_name,
                                            category=category,
                                            discipline=discipline,
                                            event=event,
                                            draw_type=bracket_draw_type
                                            )

        bracket_size = next_power_of_2(total_players)
        total_rounds = int(math.log2(bracket_size))

        if will_go_to_scoring:
            elimination_rounds = total_rounds - int(math.log2(int(finals_size)))
        else:
            elimination_rounds = total_rounds

        for round_number in range(elimination_rounds):
            round_label = total_rounds - 1 - round_number
            matches_in_round = bracket_size // (2 ** (round_number + 1))

            for match_number in range(1, matches_in_round + 1):
                Match.objects.create(
                    bracket=new_bracket,
                    round_number=round_label,
                    match_number=match_number
                )
        
        if will_go_to_scoring:

            scoring_round = ScoringRound.objects.create(
                bracket=new_bracket,
                round_number=0,
                is_final=True
            )

            if total_players > int(finals_size):
                # Link the last elimination round's matches to feed into the scoring round
                last_elim_round_label = total_rounds - elimination_rounds
                last_elim_matches = Match.objects.filter(
                    bracket=new_bracket,
                    round_number=last_elim_round_label
                )
                for match in last_elim_matches:
                    match.feeds_into_scoring = scoring_round
                    match.save()

                for i in range(int(finals_size)):
                    ScoringEntry.objects.create(scoring_round=scoring_round, entry_number=i)

            else:
                for i, reg in enumerate(registrations):
                    if discipline.is_team:
                        ScoringEntry.objects.create(scoring_round=scoring_round, team=reg.team, entry_number=i)
                    else:
                        ScoringEntry.objects.create(scoring_round=scoring_round, person=reg.person, entry_number=i)
    
        else:
            # Create 3rd place match only if there are semi-finals (4+ players)
            third_place_match = None
            if total_players >= 4:
                third_place_match = Match.objects.create(
                    bracket=new_bracket,
                    round_number=0,
                    match_number=2,
                    is_third_place=True
                )

                # Link semi-final matches to the 3rd place match so losers are advanced there
                semi_final_round = 1 
                semi_final_matches = Match.objects.filter(
                    bracket=new_bracket,
                    round_number=semi_final_round
                )
                for semi in semi_final_matches:
                    semi.loser_goes_to = third_place_match
                    semi.save()

        first_round_matches = Match.objects.filter(
            bracket=new_bracket,
            round_number=total_rounds - 1  # first round is the highest number
        ).order_by("match_number")

        real_matches_count = total_players - (bracket_size // 2)
        total_matches = len(first_round_matches)

        # Calculate evenly spread positions for real matches
        if real_matches_count > 0:
            step = total_matches / real_matches_count
            real_positions = set(int((i + 0.5) * step) for i in range(real_matches_count))
        else:
            real_positions = set()

        reg_index = 0

        for i, match in enumerate(first_round_matches):
            if i in real_positions:
                # Real match — 2 contenders
                if discipline.is_team:
                    match.team_contender_1 = registrations[reg_index].team
                else:
                    match.contender_1 = registrations[reg_index].person
                reg_index += 1
                if discipline.is_team:
                    match.team_contender_2 = registrations[reg_index].team
                else:
                    match.contender_2 = registrations[reg_index].person
                reg_index += 1
            else:
                # BYE match — 1 contender
                if discipline.is_team:
                    match.team_contender_1 = registrations[reg_index].team
                else:
                    match.contender_1 = registrations[reg_index].person
                reg_index += 1

            match.save()

        for match in first_round_matches:
            if discipline.is_team:
                if match.team_contender_1 and not match.team_contender_2:
                    match.team_winner = match.team_contender_1
                elif match.team_contender_2 and not match.team_contender_1:
                    match.team_winner = match.team_contender_2
                match.save()
                match.advance_winner()
                match.advance_loser()
            else:
                if match.contender_1 and not match.contender_2:
                    match.winner = match.contender_1
                elif match.contender_2 and not match.contender_1:
                    match.winner = match.contender_2
                match.save()
                match.advance_winner()
                match.advance_loser()
    
        return Response(
                    {"message": "Sorteio regenerado com sucesso."},
                    status=status.HTTP_200_OK
                )

    @action(
        detail=True, 
        methods=["get"], 
        url_path="export_bracket_draw", 
        permission_classes=[IsAdminRoleorHigherForGET]
    )
    def export_bracket_draw(self, request, pk=None):
        bracket = self.get_object()
        event = bracket.event
        discipline = bracket.discipline
        is_team = discipline.is_team

        dorsals = {
            ed.person_id: ed.dorsal
            for ed in EventDorsal.objects.filter(event=event)
        }

        wb = Workbook()
        ws = wb.active
        ws.title = bracket.name[:31]  # Excel sheet name limit

        if is_team:
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
        else:
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 10

        bold = Font(bold=True, name='Arial')
        normal = Font(name='Arial')
        center = Alignment(horizontal='center')
        row = 1

        if bracket.draw_type in ("Torneio/Finais", "Misto"):

            # Regular matches ordered by round (highest first = first round) then match number
            first_round_number = Match.objects.filter(
                bracket=bracket,
            ).aggregate(m=Max('round_number'))['m']

            matches = Match.objects.filter(
                bracket=bracket,
                round_number=first_round_number,
            ).select_related(
                'contender_1__club',
                'contender_2__club',
            ).order_by('match_number')

            for match in matches:
                if not discipline.is_team:
                    # Contender 1
                    if match.contender_1:
                        ws.cell(row=row, column=1, value=f'{match.contender_1.first_name} {match.contender_1.last_name}').font = normal
                        ws.cell(row=row, column=2, value=match.contender_1.club.username).font = normal
                        ws.cell(row=row, column=3, value=get_dorsal(match.contender_1, dorsals)).font = normal
                    else:
                        ws.cell(row=row, column=1, value='bye').font = normal
                    row += 1

                    # vs row
                    ws.cell(row=row, column=1, value='vs').font = bold
                    ws.cell(row=row, column=1).alignment = center
                    row += 1

                    # Contender 2
                    if match.contender_2:
                        ws.cell(row=row, column=1, value=f'{match.contender_2.first_name} {match.contender_2.last_name}').font = normal
                        ws.cell(row=row, column=2, value=match.contender_2.club.username).font = normal
                        ws.cell(row=row, column=3, value=get_dorsal(match.contender_2, dorsals)).font = normal
                    else:
                        ws.cell(row=row, column=1, value='bye').font = normal
                    row += 1
                else:
                    # Team Contender 1
                    if match.team_contender_1:
                        ws.cell(row=row, column=1, value=match.team_contender_1.club.username).font = normal
                        ws.cell(row=row, column=2, value=f'{match.team_contender_1.athlete1.first_name} {match.team_contender_1.athlete1.last_name}').font = normal
                        ws.cell(row=row, column=3, value=f'{match.team_contender_1.athlete2.first_name} {match.team_contender_1.athlete2.last_name}').font = normal
                        if match.team_contender_1.athlete3 is not None:
                            ws.cell(row=row, column=4, value=f'{match.team_contender_1.athlete3.first_name} {match.team_contender_1.athlete3.last_name}').font = normal
                        ws.cell(row=row, column=5, value=get_dorsal(match.team_contender_1.athlete1, dorsals)).font = normal
                        ws.cell(row=row, column=6, value=get_dorsal(match.team_contender_1.athlete2, dorsals)).font = normal
                        if match.team_contender_1.athlete3 is not None:
                            ws.cell(row=row, column=7, value=get_dorsal(match.team_contender_1.athlete3, dorsals)).font = normal
                        if match.team_contender_1.athlete4 is not None:
                            ws.cell(row=row, column=8, value=f'{match.team_contender_1.athlete4.first_name} {match.team_contender_1.athlete4.last_name}').font = normal
                            ws.cell(row=row, column=9, value=get_dorsal(match.team_contender_1.athlete4, dorsals)).font = normal
                    else:
                        ws.cell(row=row, column=1, value='bye').font = normal
                    row += 1

                    # vs row
                    ws.cell(row=row, column=1, value='vs').font = bold
                    ws.cell(row=row, column=1).alignment = center
                    row += 1

                    # Team Contender 2
                    if match.team_contender_2:
                        ws.cell(row=row, column=1, value=match.team_contender_2.club.username).font = normal
                        ws.cell(row=row, column=2, value=f'{match.team_contender_2.athlete1.first_name} {match.team_contender_2.athlete1.last_name}').font = normal
                        ws.cell(row=row, column=3, value=f'{match.team_contender_2.athlete2.first_name} {match.team_contender_2.athlete2.last_name}').font = normal
                        if match.team_contender_2.athlete3 is not None:
                            ws.cell(row=row, column=4, value=f'{match.team_contender_2.athlete3.first_name} {match.team_contender_2.athlete3.last_name}').font = normal
                        ws.cell(row=row, column=5, value=get_dorsal(match.team_contender_2.athlete1, dorsals)).font = normal
                        ws.cell(row=row, column=6, value=get_dorsal(match.team_contender_2.athlete2, dorsals)).font = normal
                        if match.team_contender_2.athlete3 is not None:
                            ws.cell(row=row, column=7, value=get_dorsal(match.team_contender_2.athlete3, dorsals)).font = normal
                        if match.team_contender_2.athlete4 is not None:
                            ws.cell(row=row, column=8, value=f'{match.team_contender_2.athlete4.first_name} {match.team_contender_2.athlete4.last_name}').font = normal
                            ws.cell(row=row, column=9, value=get_dorsal(match.team_contender_2.athlete4, dorsals)).font = normal
                    else:
                        ws.cell(row=row, column=1, value='bye').font = normal
                    row += 1


        scoring = ScoringRound.objects.filter(bracket=bracket).first()
        if not matches and scoring:
            entries = ScoringEntry.objects.filter(scoring_round=scoring).order_by('entry_number')
            for entry in entries:
                person = entry.person if not discipline.is_team else None
                team = entry.team if discipline.is_team else None

                if person:
                    ws.cell(row=row, column=1, value=f'{person.first_name} {person.last_name}').font = normal
                    ws.cell(row=row, column=2, value=person.club.username).font = normal
                    ws.cell(row=row, column=3, value=get_dorsal(person, dorsals)).font = normal
                elif team:
                    ws.cell(row=row, column=1, value=team.club.username).font = normal
                    ws.cell(row=row, column=2, value=f'{team.athlete1.first_name} {team.athlete1.last_name}').font = normal
                    ws.cell(row=row, column=3, value=f'{team.athlete2.first_name} {team.athlete2.last_name}').font = normal
                    if team.athlete3 is not None:
                        ws.cell(row=row, column=4, value=f'{team.athlete3.first_name} {team.athlete3.last_name}').font = normal
                    ws.cell(row=row, column=5, value=get_dorsal(team.athlete1, dorsals)).font = normal
                    ws.cell(row=row, column=6, value=get_dorsal(team.athlete2, dorsals)).font = normal
                    if team.athlete3 is not None:
                        ws.cell(row=row, column=7, value=get_dorsal(team.athlete3, dorsals)).font = normal
                else:
                    ws.cell(row=row, column=1, value='TBD').font = normal
                
                row += 1

        # Stream to response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"bracket_{bracket.name}.xlsx".replace(" ", "_")
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @action(
        detail=True,
        methods=["post"],
        url_path="officialize",
        permission_classes=[IsTechnicianOrAdminforPOST],
    )
    def officialize(self, request, pk=None):
        bracket = self.get_object()
        is_team = bracket.discipline.is_team
        third = None
        
        if bracket.draw_type == "Torneio/Finais":
            matches = Match.objects.filter(bracket=bracket).select_related(
                "winner", 
                "contender_1", 
                "contender_2", 
                "team_winner", 
                "team_contender_1", 
                "team_contender_2"
            )

            final = matches.filter(round_number=0, match_number=1).first()
            third_place_match = matches.filter(round_number=0, is_third_place=True).first()
            if is_team:
                if not final or not final.team_winner:
                    return Response({"error": "Ainda não existe um vencedor para este Escalão! Termine todas as partidas."}, status=400)

                if third_place_match and not third_place_match.team_winner:
                    return Response({"error": "Ainda não existe um terceiro classificado para este Escalão!"}, status=400)

                # Derive podium
                first = final.team_winner
                second = final.team_contender_1 if final.team_contender_2 == first else final.team_contender_2
                if third_place_match:
                    third = third_place_match.team_winner
            else:
                if not final or not final.winner:
                    return Response({"error": "Ainda não existe um vencedor para este Escalão! Termine todas as partidas."}, status=400)
                
                if third_place_match and not third_place_match.winner:
                    return Response({"error": "Ainda não existe um terceiro classificado para este Escalão!"}, status=400)

                # Derive podium
                first = final.winner
                second = final.contender_1 if final.contender_2 == first else final.contender_2
                if third_place_match:
                    third = third_place_match.winner

        elif bracket.draw_type == "Misto":
            scoring_entries = ScoringEntry.objects.filter(scoring_round__bracket=bracket).select_related("person", "team")

            rank_1 = scoring_entries.filter(rank=1).first()
            rank_2 = scoring_entries.filter(rank=2).first()
            if rank_1 is None or rank_2 is None:
                return Response({"error": "Ainda não existe um vencedor para este Escalão! Termine todas as partidas."}, status=400)

            if is_team:
                first = scoring_entries.filter(rank=1).first().team
                second = scoring_entries.filter(rank=2).first().team
                check_third = scoring_entries.filter(rank=3).first()
                if check_third is not None:
                    third = check_third.team
            else:
                first = scoring_entries.filter(rank=1).first().person
                second = scoring_entries.filter(rank=2).first().person
                check_third = scoring_entries.filter(rank=3).first()
                if check_third is not None:
                    third = check_third.person

        # reset if mistake
        Classification.objects.filter(bracket=bracket).delete()

        if is_team:
            Classification.objects.create(bracket=bracket, team=first, place=1)
            Classification.objects.create(bracket=bracket, team=second, place=2)
            if third is not None:
                Classification.objects.create(bracket=bracket, team=third, place=3)
        else:
            Classification.objects.create(bracket=bracket, person=first, place=1)
            Classification.objects.create(bracket=bracket, person=second, place=2)
            if third is not None:
                Classification.objects.create(bracket=bracket, person=third, place=3)

        serializer = ClassificationsSerializer(
            Classification.objects.filter(bracket=bracket).order_by("place"),
            many=True
        )

        bracket.officialized_at = timezone.now()
        bracket.save(update_fields=["officialized_at"])

        return Response(serializer.data)



class MatchViewSet(MultipleSerializersMixIn, viewsets.ModelViewSet):
    queryset = Match.objects.select_related(
        'bracket__event',
        'contender_1',
        'contender_2',
        'winner',
        'kataresult',
        'kumiteresult',
    ).order_by("created_at")
    serializer_class=serializers.MatchSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filterset_class = MatchesFilters
    pagination_class = None

    serializer_classes = {
        "create": serializers.CreateMatchSerializer,
        "update": serializers.UpdateMatchSerializer,
        "partial_update": serializers.UpdateMatchSerializer
    }

    def get_serializer_context(self):
        context = super().get_serializer_context()
        bracket_id = self.request.query_params.get("bracket")
        if bracket_id:
            from .models import Bracket
            bracket = Bracket.objects.select_related("event").filter(id=bracket_id).first()
            if bracket:
                event = bracket.event
                context["dorsals"] = {
                    ed.person_id: ed.dorsal
                    for ed in EventDorsal.objects.filter(event=event)
                }
        
        return context

    def perform_update(self, serializer):
        instance = serializer.save()
        if instance.ongoing and instance.is_final and instance.winner != None:
            instance.ongoing = False
            instance.save()
        if instance.ongoing:
            instance.set_ongoing()

    @action(
        detail=True,
        methods=["patch"],
        url_path="set_winner",
        permission_classes=[IsTechnicianOrAdmin],
        serializer_class=serializers.PatchMatchWinnerSerializer
    )
    def set_winner(self, request, pk=None):
        match = self.get_object()

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        winner_number = serializer.validated_data["winner"]

        try:
            match.set_winner(winner_number)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)
        
        # if final, auto goes to not ongoing
        if match.round_number == 0 and match.match_number == 1:
            match.ongoing = False
            match.save()

        if match.bracket.discipline.is_team:
            winner = match.team_winner.id
        else:
            winner = match.winner.id
        return Response({
            "success": True,
            "match_id": match.id,
            "winner": winner,
            "state": match.ongoing,
            "is_final": match.round_number == 0 and match.match_number == 1,
            "is_third_place": match.is_third_place,
            "discipline": match.bracket.discipline.name,
            "category": f"{match.bracket.category.name} {match.bracket.category.gender}"
        })

    @action(
        detail=True,
        methods=["patch"],
        url_path="advance_match",
        permission_classes=[IsTechnicianOrAdmin],
        serializer_class=serializers.AdvanceMatchSerializer
    )
    def advance_match(self, request, pk=None):
        current_match = self.get_object()

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        next_match_id = serializer.validated_data["next_match_id"]

        try:
            next_match = Match.objects.get(id=next_match_id, bracket=current_match.bracket)
        except Match.DoesNotExist:
            return Response({"error": "Match not found in this bracket."}, status=404)

        # if the provided match is a BYE, walk forward until we find one with both contenders
        ordered_matches = list(
            Match.objects.filter(bracket=current_match.bracket)
            .order_by("-round_number", "match_number")  # mirrors frontend ordering
        )

        if not next_match.contender_1 or not next_match.contender_2:
            start_index = next((i for i, m in enumerate(ordered_matches) if m.id == next_match.id), None)
            next_match = None
            if start_index is not None:
                for m in ordered_matches[start_index + 1:]:
                    if m.contender_1 and m.contender_2:
                        next_match = m
                        break

        if not next_match:
            return Response({"error": "No valid next match found."}, status=404)

        current_match.ongoing = False
        current_match.save(update_fields=["ongoing"])
        next_match.set_ongoing()

        return Response({
            "success": True,
            "next_match_id": next_match.id,
        })

    @action(
        detail=True,
        methods=["patch"],
        url_path="track_back_match",
        permission_classes=[IsTechnicianOrAdmin],
        serializer_class=serializers.PreviousMatchSerializer
    )
    def track_back_match(self, request, pk=None):
        current_match = self.get_object()

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        prev_match_id = serializer.validated_data["prev_match_id"]

        try:
            prev_match = Match.objects.get(id=prev_match_id, bracket=current_match.bracket)
        except Match.DoesNotExist:
            return Response({"error": "Match not found in this bracket."}, status=404)

        if prev_match.contender_1 == None and prev_match.contender_2 == None:
            return Response({
            "error": "Partida anterior não tem pelo menos um dos competidores definidos. Conclua as partidas das rondas anteriores!"
        }, status=404)

        current_match.ongoing = False
        current_match.save(update_fields=["ongoing"])
        prev_match.set_ongoing()

        return Response({
            "success": True,
            "prev_match_id": prev_match.id,
        })


class ScoringEntryViewSet(MultipleSerializersMixIn, viewsets.ModelViewSet):
    queryset = ScoringEntry.objects.select_related(
        'scoring_round__bracket__event',
        'scoring_round__bracket',
        "person",
        "scoring_result",
    ).order_by("created_at")
    serializer_class=serializers.ScoringEntrySerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    filterset_class = ScoringEntriesFilters
    pagination_class = None

    serializer_classes = {
        "create": serializers.CreateMatchSerializer,
        "update": serializers.UpdateScoringEntrySerializer,
        "partial_update": serializers.UpdateScoringEntrySerializer
    }

    def get_serializer_context(self):
        context = super().get_serializer_context()
        bracket_id = self.request.query_params.get("bracket")
        if bracket_id:
            from .models import Bracket
            bracket = Bracket.objects.select_related("event").filter(id=bracket_id).first()
            if bracket:
                event = bracket.event
                context["dorsals"] = {
                    ed.person_id: ed.dorsal
                    for ed in EventDorsal.objects.filter(event=event)
                }
        
        return context

    def perform_update(self, serializer):
        instance = serializer.save()
        if instance.ongoing:
            instance.set_ongoing()
